import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def fetch_product_details(product_url):
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(product_url, headers=headers)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    # --- Product name ---
    name_tag = soup.find("h1", class_="product-details__title")
    product_name = name_tag.get_text(strip=True) if name_tag else "Unknown"

    # --- Pack weight + unit price (/kg) ---
    unit_price, unit, pack_weight = None, None, None
    unit_span = soup.find("span", {"data-test": "product-details__unit-of-measurement"})
    if unit_span:
        text = unit_span.get_text(strip=True)
        pack_match = re.match(r"([\d.,]+\s*[A-Z]+)", text)
        if pack_match:
            pack_weight = pack_match.group(1)
        price_match = re.search(r"£\s*([\d.,]+)\s*/\s*([\d.,]+)\s*([A-Z]+)", text, re.IGNORECASE)
        if price_match:
            unit_price = float(price_match.group(1).replace(",", ""))
            unit = f"{price_match.group(2)} {price_match.group(3)}"

    # --- Overall price ---
    overall_price = None
    price_span = soup.find("span", class_="base-price__regular")
    if price_span:
        raw_price = price_span.get_text(strip=True)
        match = re.search(r"£\s*([\d.,]+)", raw_price)
        if match:
            overall_price = float(match.group(1).replace(",", ""))

    # --- Timestamp ---
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return {
        "name": product_name,
        "pack_weight": pack_weight,
        "overall_price": overall_price,
        "price_per_unit": unit_price,
        "unit": unit,
        "scraped_at": timestamp
    }


# --- Input and Output Files ---
input_file = "products.xlsx"
output_file = "scraped_products.xlsx"

# Load product list
products = pd.read_excel(input_file)

# Load existing history if available
if os.path.exists(output_file):
    history = pd.read_excel(output_file)
else:
    history = pd.DataFrame(columns=["name", "pack_weight", "overall_price", "price_per_unit", "unit", "scraped_at"])

results = []

for _, row in products.iterrows():
    url = row["url"]
    result = fetch_product_details(url)

    # Check history for this product
    product_history = history[history["name"] == result["name"]]
    if not product_history.empty:
        last_date_str = product_history["scraped_at"].max()
        last_date = pd.to_datetime(last_date_str)
        if datetime.now() - last_date < timedelta(days=7):
            print(f"⏩ Skipping {result['name']} (last scraped {last_date.date()})")
            continue

    results.append(result)

# Save updated history and auto-fit columns
if results:
    new_data = pd.DataFrame(results)
    updated_history = pd.concat([history, new_data], ignore_index=True)
    updated_history.to_excel(output_file, index=False)

    # Auto-fit columns using openpyxl
    wb = load_workbook(output_file)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding
    wb.save(output_file)

    print(f"✅ Added {len(results)} new records to {output_file} with auto-fit columns")
else:
    print("ℹ️ No new data to add (all products scraped within the last 7 days).")
